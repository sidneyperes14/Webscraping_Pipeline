# tratamento_dados3.py
import re
from pathlib import Path
from datetime import datetime

import pandas as pd
from openpyxl import load_workbook


# =========================
# Config
# =========================
try:
    import config
except Exception as e:
    raise RuntimeError(
        "Não consegui importar config.py. Garanta que existe e tem SHAREPOINT_DIR e DOWNLOADS_DIR."
    ) from e

SHAREPOINT_DIR: Path = Path(config.SHAREPOINT_DIR)
DOWNLOADS_DIR: Path = Path(config.DOWNLOADS_DIR)

ARQUIVO_DATA_ENTREGA_PATH = SHAREPOINT_DIR / "Arquivo - Data de entrega pedidos.xlsx"


# =========================
# Helpers
# =========================
def _normalize(s: str) -> str:
    return str(s).strip().lower().replace(" ", "")


def _find_uoou_file_xlsx(yyyymm: str) -> Path:
    """
    Procura o arquivo uoou_relatorio_YYYYMM.xlsx em DOWNLOADS_DIR.
    Se não achar exato, pega o mais recente com esse padrão.
    """
    exact = DOWNLOADS_DIR / f"uoou_relatorio_{yyyymm}.xlsx"
    if exact.exists():
        return exact

    pattern = re.compile(rf"^uoou_relatorio_{yyyymm}\.xlsx$", re.IGNORECASE)
    candidates = [p for p in DOWNLOADS_DIR.glob("uoou_relatorio_*.xlsx") if pattern.match(p.name)]
    if candidates:
        return max(candidates, key=lambda p: p.stat().st_mtime)

    raise FileNotFoundError(
        f"Não encontrei o arquivo uoou_relatorio_{yyyymm}.xlsx em: {DOWNLOADS_DIR}"
    )


def _get_single_sheet_name(path: Path) -> str:
    wb = load_workbook(path, read_only=True)
    try:
        if len(wb.sheetnames) != 1:
            raise RuntimeError(
                f"O arquivo {path.name} deveria ter 1 aba, mas encontrei: {wb.sheetnames}"
            )
        return wb.sheetnames[0]
    finally:
        wb.close()


def _read_uoou_xlsx(path: Path) -> pd.DataFrame:
    if not path.exists():
        raise FileNotFoundError(f"Arquivo Uoou não encontrado: {path}")

    sheet_name = _get_single_sheet_name(path)
    df = pd.read_excel(path, sheet_name=sheet_name)
    df.columns = [str(c).strip() if c is not None else "" for c in df.columns]
    return df


def _parse_any_date(value):
    """
    Converte diferentes formatos para datetime, quando possível.
    """
    if value is None or value == "":
        return None

    if isinstance(value, datetime):
        return value

    if hasattr(value, "year") and hasattr(value, "month") and hasattr(value, "day"):
        try:
            hour = getattr(value, "hour", 0)
            minute = getattr(value, "minute", 0)
            second = getattr(value, "second", 0)
            return datetime(value.year, value.month, value.day, hour, minute, second)
        except Exception:
            pass

    s = str(value).strip()
    if s == "":
        return None

    for fmt in (
        "%d/%m/%Y %H:%M",
        "%d/%m/%Y %H:%M:%S",
        "%d/%m/%Y",
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d %H:%M",
        "%Y-%m-%d",
        "%d-%m-%Y %H:%M:%S",
        "%d-%m-%Y %H:%M",
        "%d-%m-%Y",
    ):
        try:
            return datetime.strptime(s, fmt)
        except Exception:
            pass

    try:
        ts = pd.to_datetime(s, errors="coerce")
        if pd.isna(ts):
            return None
        return ts.to_pydatetime()
    except Exception:
        return None


def _to_float_ptbr(value):
    """
    Converte texto/valor numérico para float.
    """
    if value is None or value == "":
        return None

    if isinstance(value, (int, float)):
        return float(value)

    s = str(value).strip()
    if s == "":
        return None

    s = s.replace("R$", "").replace(" ", "")

    if "," in s and "." in s:
        s = s.replace(".", "").replace(",", ".")
    elif "," in s:
        s = s.replace(",", ".")

    try:
        return float(s)
    except Exception:
        return None


def _last_data_row(ws, col_idx: int = 1, start_row: int = 2) -> int:
    """
    Última linha com valor em uma coluna específica.
    """
    for r in range(ws.max_row, start_row - 1, -1):
        v = ws.cell(row=r, column=col_idx).value
        if v not in (None, ""):
            return r
    return start_row - 1


# =========================
# Core
# =========================
def run():
    yyyymm = datetime.now().strftime("%Y%m")
    current_year = datetime.now().year

    # 1) Ler origem
    uoou_path = _find_uoou_file_xlsx(yyyymm)
    print(f"Lendo Uoou: {uoou_path}")

    df_uoou = _read_uoou_xlsx(uoou_path)
    if df_uoou.empty:
        raise RuntimeError("O arquivo da Uoou está vazio.")

    if "Data do Pedido" not in df_uoou.columns:
        raise RuntimeError("Não achei a coluna 'Data do Pedido' no arquivo da Uoou.")

    print(f"Linhas lidas da Uoou: {len(df_uoou)}")

    # 2) Abrir destino
    if not ARQUIVO_DATA_ENTREGA_PATH.exists():
        raise FileNotFoundError(
            f"Não encontrei 'Arquivo - Data de entrega pedidos.xlsx' em: {ARQUIVO_DATA_ENTREGA_PATH}"
        )

    print(f"Abrindo: {ARQUIVO_DATA_ENTREGA_PATH}")
    wb = load_workbook(ARQUIVO_DATA_ENTREGA_PATH)
    if "Planilha1" not in wb.sheetnames:
        raise RuntimeError("A aba 'Planilha1' não existe no arquivo de destino.")

    ws = wb["Planilha1"]

    header_row = 1
    data_start_row = 2
    data_pedido_col_idx = 14  # coluna N

    # Mapear cabeçalhos do destino
    header_map = {}
    for c in range(1, ws.max_column + 1):
        val = ws.cell(row=header_row, column=c).value
        if val is not None:
            header_map[str(val).strip()] = c

    # valida se a coluna N é Data do Pedido
    header_n = ws.cell(row=header_row, column=data_pedido_col_idx).value
    if _normalize(header_n) != _normalize("Data do Pedido"):
        raise RuntimeError(
            f"A coluna N não parece ser 'Data do Pedido'. Valor encontrado: {header_n}"
        )

    # Colunas especiais
    money_cols = {
        "Valor do Pedido",
        "Valor Cobrado do Frete",
        "Valor da NF",
        "Custo do Frete",
    }
    date_cols = {
        "Data do Pedido",
        "Data do Pagamento Confirmado",
        "Data do Despacho",
        "Data Máxima Estimada de Entrega",
    }

    money_col_indexes = {header_map[name] for name in money_cols if name in header_map}
    date_col_indexes = {header_map[name] for name in date_cols if name in header_map}

    missing_money = money_cols - set(header_map.keys())
    missing_date = date_cols - set(header_map.keys())
    if missing_money:
        raise RuntimeError(f"Não achei as colunas monetárias no destino: {sorted(missing_money)}")
    if missing_date:
        raise RuntimeError(f"Não achei as colunas de data no destino: {sorted(missing_date)}")

    # 3) Excluir linhas do ano atual com base na coluna N = Data do Pedido
    last_row = _last_data_row(ws, col_idx=1, start_row=data_start_row)
    removed = 0

    for r in range(last_row, data_start_row - 1, -1):
        dt_val = ws.cell(row=r, column=data_pedido_col_idx).value
        dt = _parse_any_date(dt_val)
        if dt and dt.year == current_year:
            ws.delete_rows(r, 1)
            removed += 1

    print(f"Removidas {removed} linhas do ano atual ({current_year}) na coluna 'Data do Pedido'.")

    # 4) Inserir conteúdo novo na primeira linha livre
    last_row_after_delete = _last_data_row(ws, col_idx=1, start_row=data_start_row)
    start_append_row = last_row_after_delete + 1

    print(f"Inserindo {len(df_uoou)} linhas a partir da linha {start_append_row}...")

    # garantir mesma disposição de colunas do arquivo origem
    for i, (_, row_src) in enumerate(df_uoou.iterrows()):
        r = start_append_row + i

        for c, value in enumerate(row_src.tolist(), start=1):
            cell = ws.cell(row=r, column=c)

            if c in money_col_indexes:
                cell.value = _to_float_ptbr(value)
                cell.number_format = '"R$"#,##0.00'
            elif c in date_col_indexes:
                dt = _parse_any_date(value)
                cell.value = dt
                cell.number_format = "dd/mm/yyyy hh:mm"
            else:
                cell.value = value

    # 5) Salvar e fechar
    print("Salvando alterações...")
    wb.save(ARQUIVO_DATA_ENTREGA_PATH)
    wb.close()

    print("tratamento_dados3 concluído com sucesso.")


if __name__ == "__main__":
    run()