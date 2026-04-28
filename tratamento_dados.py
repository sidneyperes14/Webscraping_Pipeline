import re
from copy import copy
from pathlib import Path
from datetime import datetime

import pandas as pd
from openpyxl import load_workbook

from utils import esperar_arquivo_disponivel, matar_excel_zumbi, resalvar_via_excel


# =========================
# Config (usa config.py)
# =========================
try:
    import config  # deve ter SHAREPOINT_DIR e DOWNLOADS_DIR
except Exception as e:
    raise RuntimeError(
        "Não consegui importar config.py. Garanta que existe e tem SHAREPOINT_DIR e DOWNLOADS_DIR."
    ) from e

SHAREPOINT_DIR: Path = Path(config.SHAREPOINT_DIR)
DOWNLOADS_DIR: Path = Path(config.DOWNLOADS_DIR)

FECOMMERCE_PATH = SHAREPOINT_DIR / "fEcommerce.xlsx"
PRODUTOS_PATH = SHAREPOINT_DIR / "produtos.xlsx"


# =========================
# Helpers
# =========================
def _find_latest_file_by_prefix(prefix: str, yyyymm: str) -> Path:
    """
    Procura o arquivo {prefix}_{yyyymm}.xlsx em DOWNLOADS_DIR.
    Se não achar exato, pega o mais recente que combine com o prefixo.
    """
    exact = DOWNLOADS_DIR / f"{prefix}_{yyyymm}.xlsx"
    if exact.exists():
        return exact

    pattern = re.compile(rf"^{re.escape(prefix)}_{yyyymm}\.xlsx$", re.IGNORECASE)
    candidates = [p for p in DOWNLOADS_DIR.glob(f"{prefix}_*.xlsx") if pattern.match(p.name)]
    if candidates:
        return max(candidates, key=lambda p: p.stat().st_mtime)

    raise FileNotFoundError(
        f"Não encontrei o arquivo {prefix}_{yyyymm}.xlsx em: {DOWNLOADS_DIR}"
    )


def _read_olist_xlsx(path: Path) -> pd.DataFrame:
    """
    Lê a aba 'Notas fiscais' do Olist (XLSX) como texto (dtype=str),
    para evitar surpresas na ingestão.
    """
    if not path.exists():
        raise FileNotFoundError(f"Arquivo Olist não encontrado: {path}")

    df = pd.read_excel(path, sheet_name="Notas fiscais", dtype=str)
    df.columns = [str(c).strip() if c is not None else "" for c in df.columns]
    df = df.fillna("")
    return df


def _read_olist_bases_consolidadas(yyyymm: str) -> pd.DataFrame:
    """
    Lê os relatórios 1 das duas empresas e concatena em um único DataFrame.
    """
    path_1 = _find_latest_file_by_prefix("olist_relatorio", yyyymm)
    path_2 = _find_latest_file_by_prefix("olist_relatorio_empresa2", yyyymm)

    print(f"Lendo Olist empresa 1: {path_1} (aba 'Notas fiscais')")
    df_1 = _read_olist_xlsx(path_1)
    df_1["Origem Empresa"] = "Empresa 1"

    print(f"Lendo Olist empresa 2: {path_2} (aba 'Notas fiscais')")
    df_2 = _read_olist_xlsx(path_2)
    df_2["Origem Empresa"] = "Empresa 2"

    df_final = pd.concat([df_1, df_2], ignore_index=True)
    print(f"Bases Olist consolidadas: {len(df_1)} + {len(df_2)} = {len(df_final)} linhas")
    return df_final


def _read_produtos_xlsx(path: Path) -> pd.DataFrame:
    if not path.exists():
        raise FileNotFoundError(f"Não encontrei produtos.xlsx em: {path}")
    df = pd.read_excel(path, dtype=str)
    df.columns = [str(c).strip() if c is not None else "" for c in df.columns]
    df = df.fillna("")
    return df


def _normalize(s: str) -> str:
    return str(s).strip().lower().replace(" ", "")


def _ws_header_map(ws, header_row: int = 1) -> dict[str, int]:
    """
    Mapa: nome_coluna_normalizado -> indice_coluna (1-based)
    """
    m = {}
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=header_row, column=col).value
        if val is None:
            continue
        m[_normalize(val)] = col
    return m


def _parse_excel_date(value):
    """
    Converte valor de célula para datetime (quando possível).
    """
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value
    try:
        if hasattr(value, "year") and hasattr(value, "month") and hasattr(value, "day"):
            return datetime(value.year, value.month, value.day)
    except Exception:
        pass

    s = str(value).strip()
    for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d", "%d/%m/%y"):
        try:
            return datetime.strptime(s, fmt)
        except Exception:
            pass
    return None


def _parse_olist_date(value):
    """
    Data vinda do df_olist (string). Converte para date/datetime,
    para o Excel exibir corretamente com number_format.
    """
    dt = _parse_excel_date(value)
    if dt is None:
        return None
    return dt.date()


def _last_data_row(ws, col_idx=1, start_row=2) -> int:
    """
    Última linha com dado (não vazio) em uma coluna (padrão: coluna A).
    """
    for r in range(ws.max_row, start_row - 1, -1):
        v = ws.cell(row=r, column=col_idx).value
        if v not in (None, ""):
            return r
    return start_row - 1


def _copy_row_style(ws, src_row: int, dst_row: int, max_col: int):
    """
    Copia estilos da linha src_row para dst_row.
    """
    for c in range(1, max_col + 1):
        src = ws.cell(row=src_row, column=c)
        dst = ws.cell(row=dst_row, column=c)

        dst._style = copy(src._style)
        dst.number_format = src.number_format
        dst.font = copy(src.font)
        dst.fill = copy(src.fill)
        dst.border = copy(src.border)
        dst.alignment = copy(src.alignment)
        dst.protection = copy(src.protection)


def _to_float_ptbr(s: str):
    """
    Converte texto numérico (possivelmente pt-BR) para float.
    """
    if s is None:
        return None
    txt = str(s).strip()
    if txt == "" or txt == "-" or txt.lower() == "nan":
        return None

    txt = txt.replace("R$", "").replace(" ", "")

    if "," in txt and "." in txt:
        txt = txt.replace(".", "").replace(",", ".")
    elif "," in txt and "." not in txt:
        txt = txt.replace(",", ".")

    try:
        return float(txt)
    except Exception:
        return None


def _strip_leading_zeros(value):
    """
    Remove zeros à esquerda de códigos numéricos em texto.
    """
    if value is None:
        return None

    s = str(value).strip()
    if s == "":
        return ""

    if s.isdigit():
        return str(int(s))

    return s


# =========================
# Core
# =========================
def run():
    yyyymm = datetime.now().strftime("%Y%m")
    current_year = datetime.now().year
    current_month = datetime.now().month

    # 1) Ler Olist consolidado (empresa 1 + empresa 2)
    df_olist = _read_olist_bases_consolidadas(yyyymm)

    # 2) Ler Produtos (XLSX)
    print(f"Lendo Produtos: {PRODUTOS_PATH}")
    df_prod = _read_produtos_xlsx(PRODUTOS_PATH)

    # Monta mapas para lookup por descrição
    prod_cols_norm = {_normalize(c): c for c in df_prod.columns}

    desc_col = None
    for want in ("descricaoproduto", "descrição", "descricao", "itemdescricao"):
        for cnorm, corig in prod_cols_norm.items():
            if want in cnorm:
                desc_col = corig
                break
        if desc_col:
            break
    if desc_col is None:
        desc_col = df_prod.columns[0]

    def _find_prod_col(contains: str):
        for cnorm, corig in prod_cols_norm.items():
            if contains in cnorm:
                return corig
        return None

    prod_cod_col = (
        _find_prod_col("cód.produto")
        or _find_prod_col("cod.produto")
        or _find_prod_col("codproduto")
        or _find_prod_col("códproduto")
    )
    prod_pint_col = _find_prod_col("cdpintura") or _find_prod_col("pintura")

    if prod_cod_col is None:
        raise RuntimeError("Não achei a coluna 'Cód. Produto' no produtos.xlsx (por nome).")
    if prod_pint_col is None:
        raise RuntimeError("Não achei a coluna 'Cd Pintura' no produtos.xlsx (por nome).")

    prod_desc_norm = df_prod[desc_col].astype(str).str.strip().str.lower()
    map_cod = (
        pd.DataFrame({"k": prod_desc_norm, "v": df_prod[prod_cod_col]})
        .dropna(subset=["k"])
        .drop_duplicates(subset=["k"], keep="first")
        .set_index("k")["v"]
        .to_dict()
    )
    map_pint = (
        pd.DataFrame({"k": prod_desc_norm, "v": df_prod[prod_pint_col]})
        .dropna(subset=["k"])
        .drop_duplicates(subset=["k"], keep="first")
        .set_index("k")["v"]
        .to_dict()
    )

    # 3) Abrir fEcommerce e editar só Planilha1
    if not FECOMMERCE_PATH.exists():
        raise FileNotFoundError(f"Não encontrei fEcommerce.xlsx em: {FECOMMERCE_PATH}")

    print(f"Abrindo fEcommerce: {FECOMMERCE_PATH}")
    matar_excel_zumbi()
    esperar_arquivo_disponivel(FECOMMERCE_PATH, timeout_sec=180)
    wb = load_workbook(FECOMMERCE_PATH)
    if "Planilha1" not in wb.sheetnames:
        raise RuntimeError("A aba 'Planilha1' não existe no fEcommerce.xlsx")

    ws = wb["Planilha1"]

    ws.auto_filter.ref = None

    hmap = _ws_header_map(ws, header_row=1)

    def _feco_col_idx(col_name: str) -> int:
        key = _normalize(col_name)
        if key not in hmap:
            raise RuntimeError(f"Não achei a coluna '{col_name}' no cabeçalho da Planilha1.")
        return hmap[key]

    dt_col_idx = 9
    dt_header = ws.cell(row=1, column=dt_col_idx).value
    if dt_header is None or "fatur" not in str(dt_header).lower():
        key = _normalize("Dt Faturamento")
        if key in hmap:
            dt_col_idx = hmap[key]
        else:
            raise RuntimeError("Não consegui localizar a coluna 'Dt Faturamento' (coluna I ou por nome).")

    date_number_format = "dd/mm/yyyy"

    # 3.1) Excluir linhas do mês atual
    last_row = _last_data_row(ws, col_idx=1, start_row=2)
    removed = 0
    for r in range(last_row, 1, -1):
        dt_val = ws.cell(row=r, column=dt_col_idx).value
        dt = _parse_excel_date(dt_val)
        if dt and dt.month == current_month and dt.year == current_year:
            ws.delete_rows(r, 1)
            removed += 1

    print(f"Removidas {removed} linhas do mês atual ({current_year}-{current_month:02d}).")

    # 3.2) Append no final
    last_row = _last_data_row(ws, col_idx=1, start_row=2)
    start_append_row = last_row + 1
    style_row = last_row if last_row >= 2 else 2
    max_col = ws.max_column

    # ===== MAPEAMENTO (Olist -> fEcommerce) por NOME =====
    olist_to_feco = [
        ("CPF / CNPJ", "CNPJ/CPF"),
        ("Contato", "Razão Social"),
        ("Fantasia", "Fantasia"),
        ("Municipio", "Cidade"),
        ("UF", "UF"),
        ("Numero Nota", "Nº Pedido"),
        ("Data emissao", "Dt Faturamento"),
        ("Numero Nota", "Nº Nota"),
        ("Serie", "Série"),
        ("Item Descricao", "Descrição Produto"),
        ("Valor Produtos", "Valor Venda"),
        ("Item Valor", "Valor Unt. Real"),
        ("Item Quantidade", "Qtde Nota"),
        ("Item Total", "Vlr Merc. Item"),
        ("Valor Nota", "Total Nota"),
        ("Item CFOP", "CFOP"),
        ("Item Desconto", "Vlr. Desconto"),
        ("Item Frete", "Vlr. Frete"),
        ("Item Outras Despesas", "Vlr. Outras desp."),
    ]

    idx_desc_prod = _feco_col_idx("Descrição Produto")
    idx_cod_prod = _feco_col_idx("Cód. Produto")
    idx_cd_pintura = _feco_col_idx("Cd Pintura")
    idx_ano = _feco_col_idx("Ano")

    idx_num_pedido = _feco_col_idx("Nº Pedido")
    idx_num_nota = _feco_col_idx("Nº Nota")

    idx_vlr_merc = _feco_col_idx("Vlr Merc. Item")
    idx_vlr_desc = _feco_col_idx("Vlr. Desconto")
    idx_vlr_frete = _feco_col_idx("Vlr. Frete")
    idx_vlr_unit_calc = _feco_col_idx("Vlr. Unit. + IPI + ST + Frete + Seguro + Out. Desp. - Desconto")

    money_col_names = [
        "Valor Venda",
        "Valor Unt. Real",
        "Vlr Merc. Item",
        "Vlr. Desconto",
        "Vlr. Frete",
        "Vlr. Outras desp.",
        "Total Nota",
        "Vlr. Unit. + IPI + ST + Frete + Seguro + Out. Desp. - Desconto",
    ]
    money_cols_idx = [_feco_col_idx(name) for name in money_col_names]
    money_number_format = "#,##0.00"

    olist_cols_norm = {_normalize(c): c for c in df_olist.columns}

    def _olist_series(col_name: str):
        key = _normalize(col_name)
        if key not in olist_cols_norm:
            raise RuntimeError(f"Não achei a coluna '{col_name}' no Olist consolidado (aba Notas fiscais).")
        return df_olist[olist_cols_norm[key]]

    olist_series_map = {src: _olist_series(src) for src, _ in olist_to_feco if src}

    print(f"Inserindo {len(df_olist)} linhas novas no final da Planilha1...")

    for i in range(len(df_olist)):
        r = start_append_row + i

        _copy_row_style(ws, style_row, r, max_col)

        for src_col, dst_col in olist_to_feco:
            dst_idx = _feco_col_idx(dst_col)
            val = olist_series_map[src_col].iloc[i]

            if dst_idx == dt_col_idx:
                dt = _parse_olist_date(val)
                ws.cell(row=r, column=dst_idx).value = dt
                ws.cell(row=r, column=dst_idx).number_format = date_number_format
                continue

            if dst_idx in (idx_num_pedido, idx_num_nota):
                ws.cell(row=r, column=dst_idx).value = _strip_leading_zeros(val)
                continue

            if dst_idx in money_cols_idx:
                num = _to_float_ptbr(val)
                ws.cell(row=r, column=dst_idx).value = num
                ws.cell(row=r, column=dst_idx).number_format = money_number_format
            else:
                ws.cell(row=r, column=dst_idx).value = val

        ws.cell(row=r, column=idx_ano).value = current_year

        desc = ws.cell(row=r, column=idx_desc_prod).value
        desc_norm = str(desc).strip().lower() if desc is not None else ""
        ws.cell(row=r, column=idx_cod_prod).value = map_cod.get(desc_norm, None)
        ws.cell(row=r, column=idx_cd_pintura).value = map_pint.get(desc_norm, None)

        vlr_merc = ws.cell(row=r, column=idx_vlr_merc).value
        vlr_desc = ws.cell(row=r, column=idx_vlr_desc).value
        vlr_frete = ws.cell(row=r, column=idx_vlr_frete).value

        vlr_merc = float(vlr_merc) if vlr_merc not in (None, "") else 0.0
        vlr_desc = float(vlr_desc) if vlr_desc not in (None, "") else 0.0
        vlr_frete = float(vlr_frete) if vlr_frete not in (None, "") else 0.0

        vlr_unit_calc = vlr_merc - vlr_desc + vlr_frete
        ws.cell(row=r, column=idx_vlr_unit_calc).value = vlr_unit_calc
        ws.cell(row=r, column=idx_vlr_unit_calc).number_format = money_number_format

    print("Salvando alterações no fEcommerce.xlsx...")
    wb.save(FECOMMERCE_PATH)
    wb.close()

    print("Re-salvando fEcommerce.xlsx via Excel para normalizar o arquivo...")
    esperar_arquivo_disponivel(FECOMMERCE_PATH, timeout_sec=120)
    resalvar_via_excel(FECOMMERCE_PATH)

    print("Tratamento concluído e fEcommerce atualizado com sucesso.")


if __name__ == "__main__":
    run()