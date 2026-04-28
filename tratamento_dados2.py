# tratamento_dados2.py
import re
import unicodedata
from copy import copy
from pathlib import Path
from datetime import datetime

import pandas as pd
from openpyxl import load_workbook

from utils import esperar_arquivo_disponivel, matar_excel_zumbi, resalvar_via_excel


# =========================
# Config
# =========================
try:
    import config
except Exception as e:
    raise RuntimeError(
        "Nao consegui importar config.py. Garanta que existe e tem SHAREPOINT_DIR e DOWNLOADS_DIR."
    ) from e

SHAREPOINT_DIR: Path = Path(config.SHAREPOINT_DIR)
DASHBOARD_DATA_DIR: Path = Path(config.DASHBOARD_DATA_DIR)
DOWNLOADS_DIR: Path = Path(config.DOWNLOADS_DIR)

RESUMO_PROFORMA_PATH = DASHBOARD_DATA_DIR / "resumo_proforma_pedido.xlsx"


# =========================
# Helpers
# =========================
def _normalize(s: str) -> str:
    return str(s).strip().lower().replace(" ", "")


def _normalize_text_upper(s) -> str:
    """
    Normaliza texto para comparacoes:
    - tira acentos
    - remove espacos extras
    - converte para maiusculo
    """
    if s is None:
        return ""
    txt = str(s).strip()
    txt = unicodedata.normalize("NFKD", txt).encode("ascii", "ignore").decode("ascii")
    txt = " ".join(txt.split())
    return txt.upper()


def _find_olist2_file_xlsx(yyyymm: str) -> Path:
    """
    Procura o arquivo olist_relatorio2_YYYYMM.xlsx em DOWNLOADS_DIR.
    Se nao achar exato, pega o mais recente com esse padrao.
    """
    exact = DOWNLOADS_DIR / f"olist_relatorio2_{yyyymm}.xlsx"
    if exact.exists():
        return exact

    pattern = re.compile(rf"^olist_relatorio2_{yyyymm}\.xlsx$", re.IGNORECASE)
    candidates = [p for p in DOWNLOADS_DIR.glob("olist_relatorio2_*.xlsx") if pattern.match(p.name)]
    if candidates:
        return max(candidates, key=lambda p: p.stat().st_mtime)

    raise FileNotFoundError(
        f"Nao encontrei o arquivo olist_relatorio2_{yyyymm}.xlsx em: {DOWNLOADS_DIR}"
    )


def _read_olist2_xlsx(path: Path) -> pd.DataFrame:
    if not path.exists():
        raise FileNotFoundError(f"Arquivo nao encontrado: {path}")

    df = pd.read_excel(path, sheet_name="Sheet1", dtype=str)
    df.columns = [str(c).strip() if c is not None else "" for c in df.columns]
    df = df.fillna("")
    return df


def _ws_header_map(ws, header_row: int = 3) -> dict[str, int]:
    """
    Mapa: nome_coluna_normalizado -> indice_coluna (1-based)
    """
    m = {}
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=header_row, column=col).value
        if val is None:
            continue
        m[_normalize(val)] = col
    return m


def _last_data_row(ws, col_idx: int, start_row: int) -> int:
    """
    Ultima linha com valor em uma coluna especifica.
    """
    for r in range(ws.max_row, start_row - 1, -1):
        v = ws.cell(row=r, column=col_idx).value
        if v not in (None, ""):
            return r
    return start_row - 1


def _copy_style_only(ws, src_row: int, dst_row: int, col_idx: int):
    src = ws.cell(row=src_row, column=col_idx)
    dst = ws.cell(row=dst_row, column=col_idx)

    dst._style = copy(src._style)
    dst.number_format = src.number_format
    dst.font = copy(src.font)
    dst.fill = copy(src.fill)
    dst.border = copy(src.border)
    dst.alignment = copy(src.alignment)
    dst.protection = copy(src.protection)


def _strip_leading_zeros(value):
    """
    Remove zeros a esquerda de textos numericos.
    Exemplos:
    000015 -> 15
    000000 -> 0
    """
    if value is None:
        return None

    s = str(value).strip()
    if s == "":
        return ""

    if s.isdigit():
        return str(int(s))

    return s


def _calcular_canal(cliente, razao_social) -> str:
    """
    Reproduz em Python a logica da formula da coluna I:

    =SE(G="Uoou Solutions";"SITE";
      SE(G="Amazon";"AMAZON";
      SE(G="Mercado Livre";"MERCADO LIVRE";
      SE(H="WHATSAPP";"WHATS";
      SE(H="PARCERIAS";"PARCERIA";
      SE(H="Uoou Solutions";"SITE";"SITE"))))))
    """
    g = _normalize_text_upper(cliente)
    h = _normalize_text_upper(razao_social)

    if g == "UOOU SOLUTIONS":
        return "SITE"
    if g == "AMAZON":
        return "AMAZON"
    if g == "MERCADO LIVRE":
        return "MERCADO LIVRE"
    if h == "WHATSAPP":
        return "WHATS"
    if h == "PARCERIAS":
        return "PARCERIA"
    if h == "UOOU SOLUTIONS":
        return "SITE"

    return "SITE"


def _canal_para_codigo(canal):
    mapa = {
        "SITE": 1,
        "WHATS": 2,
        "DIRETO FABRICA": 3,
        "PARCERIA": 4,
        "MADEIRA MADEIRA": 5,
        "PJ": 6,
        "AMAZON": 7,
        "MERCADO LIVRE": 8,
    }

    if canal is None:
        return None

    chave = _normalize_text_upper(canal)
    return mapa.get(chave, 0)


# =========================
# Core
# =========================
def run():
    yyyymm = datetime.now().strftime("%Y%m")

    # 1) Ler arquivo origem
    olist2_path = _find_olist2_file_xlsx(yyyymm)
    print(f"Lendo Olist 2: {olist2_path} (aba 'Sheet1')")
    df = _read_olist2_xlsx(olist2_path)

    required_cols = ["Número (Nota Fiscal)", "E-commerce", "Intermediador"]
    for col in required_cols:
        if col not in df.columns:
            raise RuntimeError(f"Nao achei a coluna '{col}' na aba Sheet1.")

    # filtrar somente o que tem Numero (Nota Fiscal)
    df["Número (Nota Fiscal)"] = df["Número (Nota Fiscal)"].astype(str).fillna("").str.strip()
    df["E-commerce"] = df["E-commerce"].astype(str).fillna("").str.strip()
    df["Intermediador"] = df["Intermediador"].astype(str).fillna("").str.strip()

    df_filtrado = df[df["Número (Nota Fiscal)"] != ""].copy()
    print(f"Linhas validas apos filtro de 'Numero (Nota Fiscal)': {len(df_filtrado)}")

    # 2) Abrir arquivo destino
    if not RESUMO_PROFORMA_PATH.exists():
        raise FileNotFoundError(f"Nao encontrei resumo_proforma_pedido.xlsx em: {RESUMO_PROFORMA_PATH}")

    print(f"Abrindo: {RESUMO_PROFORMA_PATH}")
    matar_excel_zumbi()
    esperar_arquivo_disponivel(RESUMO_PROFORMA_PATH, timeout_sec=180)
    wb = load_workbook(RESUMO_PROFORMA_PATH)
    if "resumo_proforma_pedido" not in wb.sheetnames:
        raise RuntimeError("A aba 'resumo_proforma_pedido' nao existe no arquivo de destino.")

    ws = wb["resumo_proforma_pedido"]

    header_row = 3
    data_start_row = 4

    hmap = _ws_header_map(ws, header_row=header_row)

    def _col_idx(col_name: str) -> int:
        key = _normalize(col_name)
        if key not in hmap:
            raise RuntimeError(f"Nao achei a coluna '{col_name}' no cabecalho da linha {header_row}.")
        return hmap[key]

    idx_pedido = _col_idx("Pedido")
    idx_cliente = _col_idx("Cliente")
    idx_razao_social = _col_idx("Razão Social")
    idx_col_i = 9   # I
    idx_col_j = 10  # J

    # ultima linha preenchida na coluna Pedido
    last_row = _last_data_row(ws, col_idx=idx_pedido, start_row=data_start_row)
    start_append_row = last_row + 1

    print(f"Inserindo dados a partir da linha {start_append_row}...")

    # 3) Preencher dados novos
    for i, (_, row_src) in enumerate(df_filtrado.iterrows()):
        r = start_append_row + i

        pedido = _strip_leading_zeros(row_src["Número (Nota Fiscal)"])
        cliente = row_src["E-commerce"]
        razao_social = row_src["Intermediador"]

        ws.cell(row=r, column=idx_pedido).value = pedido
        ws.cell(row=r, column=idx_cliente).value = cliente
        ws.cell(row=r, column=idx_razao_social).value = razao_social

    # 4) Remover duplicados pela coluna Pedido, mantendo a primeira ocorrencia
    last_row_after_insert = _last_data_row(ws, col_idx=idx_pedido, start_row=data_start_row)
    seen = set()
    rows_to_delete = []

    for r in range(data_start_row, last_row_after_insert + 1):
        pedido_val = ws.cell(row=r, column=idx_pedido).value
        pedido_key = str(pedido_val).strip() if pedido_val is not None else ""

        if pedido_key == "":
            continue

        if pedido_key in seen:
            rows_to_delete.append(r)
        else:
            seen.add(pedido_key)

    for r in reversed(rows_to_delete):
        ws.delete_rows(r, 1)

    print(f"Removidas {len(rows_to_delete)} linhas duplicadas com base na coluna 'Pedido'.")

    # 5) Preencher colunas I e J por Python
    last_row_after_dedup = _last_data_row(ws, col_idx=idx_pedido, start_row=data_start_row)

    if last_row_after_dedup >= data_start_row:
        # usa a primeira linha de dados como base de estilo para I e J
        style_base_row = data_start_row

        for r in range(data_start_row, last_row_after_dedup + 1):
            if r != style_base_row:
                _copy_style_only(ws, style_base_row, r, idx_col_i)
                _copy_style_only(ws, style_base_row, r, idx_col_j)

            cliente = ws.cell(row=r, column=idx_cliente).value
            razao_social = ws.cell(row=r, column=idx_razao_social).value

            canal = _calcular_canal(cliente, razao_social)
            codigo = _canal_para_codigo(canal)

            ws.cell(row=r, column=idx_col_i).value = canal
            ws.cell(row=r, column=idx_col_j).value = codigo

        print(f"Coluna I preenchida com o canal e coluna J com o codigo ate a linha {last_row_after_dedup}.")

    # 6) Salvar e fechar
    print("Salvando alteracoes...")
    wb.save(RESUMO_PROFORMA_PATH)
    wb.close()

    print("Re-salvando resumo_proforma_pedido.xlsx via Excel para normalizar o arquivo...")
    esperar_arquivo_disponivel(RESUMO_PROFORMA_PATH, timeout_sec=120)
    resalvar_via_excel(RESUMO_PROFORMA_PATH)

    print("tratamento_dados2 concluido com sucesso.")


if __name__ == "__main__":
    run()